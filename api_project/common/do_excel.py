# -*- coding: utf-8 -*-
# Created by xj on 2019/4/8
from openpyxl import workbook,load_workbook
from api_project.common.real_path import test_data_path
from api_project.common.my_config import MyConfig


class DoExcel:
    def __init__(self,file_path,sheet_name):
        self.file_path=file_path #初始化打开用例地路径
        self.sheet_name=sheet_name #初始化用例表单名称
    def update_creid(self):
        wb=load_workbook(self.file_path)
        sheet=wb['creid']
        old_creid=sheet.cell(1,2).value
        return old_creid
    def write_back_creid(self,value):
        wb=load_workbook(self.file_path)
        sheet=wb['creid']
        sheet.cell(row=1,column=2).value=value
        wb.save(self.file_path)
        wb.close()


    def do_excel(self,SECTION):
        '读取excel中地数据'
        button=MyConfig().get_list(SECTION,'button') #用来获取配置文件中的button来控制用例选择性执行
        wb=load_workbook(self.file_path) #打开excel
        sheet=wb[self.sheet_name] #定位表单
        old_creid=self.update_creid()
        test_data=[] #定义一个大列表存放每行数据
        for i in range(2,sheet.max_row+1):
            row_data={} #创建空字典，目的是将每行数据作为一个字典储存
            row_data['Case_id']=sheet.cell(row=i,column=1).value
            row_data['Moudle']=sheet.cell(row=i,column=2).value
            row_data['Title']=sheet.cell(row=i,column=3).value
            row_data['Url']=sheet.cell(row=i,column=4).value
            row_data['Sql']=sheet.cell(row=i,column=5).value
            if sheet.cell(row=i,column=6).value.find('CREID') !=-1:
                params=sheet.cell(row=i,column=6).value.replace('CREID',str(old_creid))
                new_creid=old_creid +1
                self.write_back_creid(new_creid)
            else:
                params=sheet.cell(row=i,column=6).value
            row_data['Params']=params
            row_data['Expect_Result']=sheet.cell(row=i,column=7).value
            test_data.append(row_data)
        wb.close()
        final_data=[]
        #if else 用来控制执行测试用例的逻辑
        if button == 'ALL':
            final_data=test_data
        else:
            for i in button:
                final_data.append(test_data[i-1]) #根据索引关系来获取最终的
        return final_data #返回最后地大列表，存放所有收集到地数据


    def write_back(self,i,j,value):
        '定义一个excel写回函数'
        wb=load_workbook(self.file_path)
        sheet=wb[self.sheet_name]
        sheet.cell(row=i,column=j).value=value #写回excel数据
        wb.save(self.file_path)
        wb.close()
    def creat_sheet(self):
        '定义一个创建表单的函数'
        wb=workbook.Workbook()
        wb.create_sheet(self.sheet_name)
        wb.save(self.file_path)
        wb.close()
if __name__ == '__main__':
    test_data=DoExcel(test_data_path,'verifiedUserAuth').do_excel('VERIFYUSERAUTH')
    print(test_data)






